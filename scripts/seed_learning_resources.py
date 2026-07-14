"""
scripts/seed_learning_resources.py

One-time (idempotent) seed script that loads the curated learning-resource
catalogs into the `learning_resources` table the Recommendation Engine reads
from (see `services/recommendation_engine.py`). Safe to re-run: matches on
`url` (the table's unique constraint, see migration 0003) and skips rows
that already exist rather than duplicating or overwriting them.

Requires migration 0003 (`alembic upgrade head`) to have already created
`learning_resources` — this script only inserts rows, it never creates
tables.

Usage:
    python -m scripts.seed_learning_resources
    python -m scripts.seed_learning_resources --vietfuture path/to/VietFuture.xlsx --ted path/to/TED.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
import pandas as pd
from sqlalchemy.orm import Session as DBSession

from db.models import LearningResourceORM
from db.session import SessionLocal
from utils.logger import get_logger

logger = get_logger(__name__)

_DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "learning_resources"
_DEFAULT_VIETFUTURE = _DATA_DIR / "VietFuture.xlsx"
_DEFAULT_TED = _DATA_DIR / "Danh_Sach_TED_Talk_Ky_Nang.xlsx"

# Maps each catalog's raw Vietnamese category text to the normalized
# skill-tag slugs the Recommendation Engine matches against a session's weak
# sub-scores (see services/recommendation_engine.py::SKILL_TAG_TO_SCORE_FIELDS).
# Both source catalogs use overlapping-but-not-identical category labels
# (e.g. "Kỹ năng nói" vs. "Kỹ năng nói & Ngôn ngữ cơ thể"), so this maps by
# exact label rather than assuming the two files share a taxonomy.
_CATEGORY_TO_SKILL_TAGS: dict[str, list[str]] = {
    "Cải thiện độ tự tin": ["confidence"],
    "Kỹ năng nói": ["speaking"],
    "Kỹ năng nói & Ngôn ngữ cơ thể": ["speaking"],
    "Kỹ năng thuyết trình": ["presentation"],
    "Kỹ năng phản biện": ["critical_thinking"],
    "Kỹ năng phỏng vấn": ["interview"],
    "Cải thiện kỹ năng": ["general"],
}


def _skill_tags_for(category: str | None) -> list[str]:
    if not category:
        return ["general"]
    return _CATEGORY_TO_SKILL_TAGS.get(category.strip(), ["general"])


def _clean(value: object) -> str | None:
    text = str(value).strip() if value is not None else ""
    if not text or text.lower() == "nan":
        return None
    return text


def load_vietfuture(path: Path) -> list[dict]:
    """
    Parses VietFuture.xlsx.

    Columns (see data/learning_resources/VietFuture.xlsx): `Catergory` [sic],
    `Title` (often "Title | Speaker | Channel", pipe-separated -- not every
    row has all three parts), `Link`, `Platform`, `Langauge ` [sic, trailing
    space]. Plain-text URLs (no hyperlink-only cells here, unlike the TED
    catalog), so a straightforward `pandas.read_excel` is sufficient.
    """
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]

    rows = []
    for _, row in df.iterrows():
        url = _clean(row.get("Link"))
        if not url:
            continue

        raw_title = _clean(row.get("Title")) or ""
        parts = [p.strip() for p in raw_title.split("|")]
        title = parts[0] if parts and parts[0] else raw_title
        speaker = parts[1] if len(parts) > 1 else None
        source = parts[2] if len(parts) > 2 else None

        category = _clean(row.get("Catergory"))
        platform = _clean(row.get("Platform"))
        language_raw = (_clean(row.get("Langauge")) or "").upper()
        language = "vi" if language_raw in ("VN", "VI") else "en" if language_raw else None

        rows.append(
            {
                "title": title,
                "url": url,
                "resource_type": "video" if platform and "youtube" in platform.lower() else "article",
                "platform": platform,
                "language": language,
                "speaker": speaker,
                "source": source,
                "description": None,
                "skill_tags": _skill_tags_for(category),
                "category_label": category,
            }
        )
    return rows


def load_ted_talks(path: Path) -> list[dict]:
    """
    Parses Danh_Sach_TED_Talk_Ky_Nang.xlsx.

    The real header row is row 4 (1-indexed) -- rows 1-3 are a title/blank
    row, per the sheet's own layout. Columns: `STT`, `Danh Mục`,
    `Tiêu Đề Bài Nói (TED Talk)`, `Diễn Giả`, `Đường Dẫn (Link YouTube)`
    (the cell's displayed text is just "Xem trên YouTube" -- the real URL is
    a hyperlink attached to the cell, so this must be read via openpyxl's
    `cell.hyperlink.target`, not a plain-text/pandas read), and
    `Nội Dung Chính & Giá Trị Cốt Lõi`.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_row = 4
    headers = [cell.value for cell in ws[header_row]]
    col_index = {name: idx for idx, name in enumerate(headers) if name is not None}

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1):
        if row[col_index["STT"]].value is None:
            continue

        link_cell = row[col_index["Đường Dẫn (Link YouTube)"]]
        url = link_cell.hyperlink.target if link_cell.hyperlink else _clean(link_cell.value)
        if not url:
            continue

        category = _clean(row[col_index["Danh Mục"]].value)
        rows.append(
            {
                "title": _clean(row[col_index["Tiêu Đề Bài Nói (TED Talk)"]].value) or "",
                "url": url,
                "resource_type": "video",
                "platform": "Youtube",
                "language": "en",
                "speaker": _clean(row[col_index["Diễn Giả"]].value),
                "source": "TED Talk",
                "description": _clean(row[col_index["Nội Dung Chính & Giá Trị Cốt Lõi"]].value),
                "skill_tags": _skill_tags_for(category),
                "category_label": category,
            }
        )
    return rows


def seed(db: DBSession, vietfuture_path: Path, ted_path: Path) -> tuple[int, int]:
    """Inserts every not-yet-seen resource (matched by URL). Returns (inserted, skipped)."""
    existing_urls = {url for (url,) in db.query(LearningResourceORM.url).all()}

    all_rows: list[dict] = []
    if vietfuture_path.exists():
        all_rows.extend(load_vietfuture(vietfuture_path))
    else:
        logger.warning("VietFuture catalog not found at %s, skipping", vietfuture_path)
    if ted_path.exists():
        all_rows.extend(load_ted_talks(ted_path))
    else:
        logger.warning("TED Talk catalog not found at %s, skipping", ted_path)

    inserted = skipped = 0
    for row in all_rows:
        if not row["title"] or row["url"] in existing_urls:
            skipped += 1
            continue
        db.add(LearningResourceORM(**row))
        existing_urls.add(row["url"])
        inserted += 1

    db.commit()
    return inserted, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--vietfuture", type=Path, default=_DEFAULT_VIETFUTURE)
    parser.add_argument("--ted", type=Path, default=_DEFAULT_TED)
    args = parser.parse_args()

    db = SessionLocal()
    try:
        inserted, skipped = seed(db, args.vietfuture, args.ted)
        logger.info("Seeded %d learning resources (%d already present, skipped).", inserted, skipped)
        print(f"Inserted {inserted} learning resources, skipped {skipped} already-present rows.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
